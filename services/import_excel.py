"""
Importación de datos maestros desde Excel (.xlsx / .xlsm).

Usa openpyxl directamente (sin pandas). Valida el archivo antes de importar.
Crea backup automático antes de cada importación.
Informa de registros nuevos, duplicados saltados y errores por fila.
"""

import os
from services.backup_manager import crear_backup

# Nombres de hojas y columnas esperadas en el Excel legacy
HOJA_BASES = 'BASES'
HOJA_GRANJAS = 'GRANJAS'


def validar_excel(ruta_excel):
    """
    Valida que el archivo Excel existe, se puede abrir y tiene las hojas esperadas.

    Returns:
        dict: {
            'valido': bool,
            'mensaje': str,
            'hojas_encontradas': list,
            'hojas_faltantes': list
        }
    """
    resultado = {
        'valido': False,
        'mensaje': '',
        'hojas_encontradas': [],
        'hojas_faltantes': []
    }

    if not ruta_excel:
        resultado['mensaje'] = "No se ha indicado la ruta al archivo Excel en la configuración."
        return resultado

    if not os.path.exists(ruta_excel):
        resultado['mensaje'] = (
            f"El archivo Excel no se encuentra en la ruta indicada:\n{ruta_excel}\n\n"
            f"Compruebe que la ruta es correcta en la configuración."
        )
        return resultado

    ext = os.path.splitext(ruta_excel)[1].lower()
    if ext not in ('.xlsx', '.xlsm'):
        resultado['mensaje'] = (
            f"El archivo debe tener extensión .xlsx o .xlsm.\n"
            f"El archivo actual tiene extensión '{ext}'."
        )
        return resultado

    try:
        import openpyxl
        wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
        hojas = wb.sheetnames
        wb.close()
    except Exception as e:
        resultado['mensaje'] = (
            f"No se puede abrir el archivo Excel.\n"
            f"Compruebe que no está abierto en otro programa.\n"
            f"Detalle técnico: {e}"
        )
        return resultado

    resultado['hojas_encontradas'] = hojas

    hojas_requeridas = [HOJA_BASES]
    faltantes = [h for h in hojas_requeridas if h not in hojas]
    resultado['hojas_faltantes'] = faltantes

    if faltantes:
        resultado['mensaje'] = (
            f"El Excel no tiene las hojas necesarias.\n"
            f"Se necesita la hoja: {', '.join(faltantes)}.\n"
            f"Hojas encontradas: {', '.join(hojas)}."
        )
        return resultado

    resultado['valido'] = True
    resultado['mensaje'] = "Excel válido y listo para importar."
    return resultado


def importar_maestros(ruta_excel):
    """
    Importa vehículos, conductores y granjas desde el Excel legacy.

    - Valida antes de tocar la DB.
    - Crea backup automático.
    - Salta registros duplicados (ya existentes por código).
    - Informa de nuevos, saltados y errores por fila.

    Returns:
        dict: Resumen detallado de la importación.
    """
    import openpyxl
    from db.connection import get_db

    resultado = {
        'vehiculos_nuevos': 0,
        'vehiculos_saltados': 0,
        'conductores_nuevos': 0,
        'conductores_saltados': 0,
        'granjas_nuevas': 0,
        'granjas_saltadas': 0,
        'errores': [],
        'avisos': [],
        'exito': False
    }

    # Validar primero
    validacion = validar_excel(ruta_excel)
    if not validacion['valido']:
        resultado['errores'].append(validacion['mensaje'])
        return resultado

    # Backup antes de importar
    backup_path = crear_backup(motivo='pre_import')
    if backup_path:
        resultado['avisos'].append("Copia de seguridad creada automáticamente.")

    try:
        wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
    except Exception as e:
        resultado['errores'].append(f"Error abriendo Excel: {e}")
        return resultado

    db = get_db()

    # ─── VEHÍCULOS (hoja BASES, columnas A:C) ───────────────────
    try:
        if HOJA_BASES in wb.sheetnames:
            ws = wb[HOJA_BASES]
            for i, row in enumerate(ws.iter_rows(min_row=3, max_col=3, values_only=True), start=3):
                codigo = row[0]
                if codigo is None:
                    continue
                codigo = str(codigo).split('.')[0].strip()
                if not codigo:
                    continue

                # Verificar si ya existe
                existente = db.execute(
                    "SELECT id FROM vehiculos WHERE codigo_interno = ?", (codigo,)
                ).fetchone()
                if existente:
                    resultado['vehiculos_saltados'] += 1
                    continue

                matricula_t = str(row[1]).strip() if row[1] else ''
                matricula_s = str(row[2]).strip() if row[2] else ''
                try:
                    db.execute(
                        "INSERT INTO vehiculos (codigo_interno, matricula_tractora, matricula_semirremolque) "
                        "VALUES (?, ?, ?)",
                        (codigo, matricula_t, matricula_s)
                    )
                    resultado['vehiculos_nuevos'] += 1
                except Exception as e:
                    resultado['errores'].append(f"Fila {i} (vehículo '{codigo}'): {e}")
            db.commit()
    except Exception as e:
        resultado['errores'].append(f"Error leyendo vehículos del Excel: {e}")

    # ─── CONDUCTORES (hoja BASES, columnas G:I) ─────────────────
    try:
        if HOJA_BASES in wb.sheetnames:
            ws = wb[HOJA_BASES]
            for i, row in enumerate(ws.iter_rows(min_row=3, min_col=7, max_col=9, values_only=True), start=3):
                alias = str(row[0]).strip() if row[0] else ''
                if not alias:
                    continue

                existente = db.execute(
                    "SELECT id FROM conductores WHERE alias = ?", (alias,)
                ).fetchone()
                if existente:
                    resultado['conductores_saltados'] += 1
                    continue

                cod_alfa = str(row[1]).strip() if row[1] else ''
                dni = str(row[2]).strip() if row[2] else ''
                try:
                    db.execute(
                        "INSERT INTO conductores (alias, codigo_alfabetico, dni) VALUES (?, ?, ?)",
                        (alias, cod_alfa, dni)
                    )
                    resultado['conductores_nuevos'] += 1
                except Exception as e:
                    resultado['errores'].append(f"Fila {i} (conductor '{alias}'): {e}")
            db.commit()
    except Exception as e:
        resultado['errores'].append(f"Error leyendo conductores del Excel: {e}")

    # ─── GRANJAS (hoja GRANJAS) ──────────────────────────────────
    try:
        if HOJA_GRANJAS in wb.sheetnames:
            ws = wb[HOJA_GRANJAS]
            for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # Columna B = código (índice 1)
                if len(row) < 3:
                    continue
                codigo = str(row[1]).split('.')[0].strip() if row[1] else ''
                if not codigo:
                    continue

                existente = db.execute(
                    "SELECT id FROM granjas WHERE codigo = ?", (codigo,)
                ).fetchone()
                if existente:
                    resultado['granjas_saltadas'] += 1
                    continue

                nombre = str(row[2]).strip() if row[2] else 'DESCONOCIDO'
                localidad = str(row[3]).strip() if len(row) > 3 and row[3] else ''

                trayecto = 120
                carga = 60
                try:
                    if len(row) > 6 and row[6] is not None:
                        trayecto = int(row[6])
                except (ValueError, TypeError):
                    pass
                try:
                    if len(row) > 7 and row[7] is not None:
                        carga = int(row[7])
                except (ValueError, TypeError):
                    pass

                try:
                    db.execute(
                        "INSERT INTO granjas (codigo, nombre_cliente, localidad, "
                        "tiempo_trayecto_min, tiempo_carga_min) VALUES (?, ?, ?, ?, ?)",
                        (codigo, nombre, localidad, trayecto, carga)
                    )
                    resultado['granjas_nuevas'] += 1
                except Exception as e:
                    resultado['errores'].append(f"Fila {i} (granja '{codigo}'): {e}")
            db.commit()
        else:
            resultado['avisos'].append(
                f"La hoja '{HOJA_GRANJAS}' no existe en el Excel. "
                f"Las granjas no se importaron."
            )
    except Exception as e:
        resultado['errores'].append(f"Error leyendo granjas del Excel: {e}")

    wb.close()

    # Añadir avisos sobre saltados
    total_saltados = (resultado['vehiculos_saltados'] +
                      resultado['conductores_saltados'] +
                      resultado['granjas_saltadas'])
    if total_saltados > 0:
        resultado['avisos'].append(
            f"Se saltaron {total_saltados} registros que ya existían en la base de datos "
            f"({resultado['vehiculos_saltados']} vehículos, "
            f"{resultado['conductores_saltados']} conductores, "
            f"{resultado['granjas_saltadas']} granjas)."
        )

    resultado['exito'] = len(resultado['errores']) == 0
    return resultado
